import csv
import io
import os
from datetime import date

import asyncpg
from openpyxl import load_workbook

from app.models.parcial import Parcial
from app.models.importacion_archivo import ImportacionArchivo
from app.repositories.carrera_repository import CarreraRepository
from app.repositories.crud_repository import CrudRepository, CrudTableConfig
from app.repositories.cursada_repository import CursadaRepository
from app.repositories.estudiante_repository import EstudianteRepository
from app.repositories.importacion_archivo_repository import ImportacionArchivoRepository
from app.repositories.inscripcion_cuatrimestre_repository import InscripcionCuatrimestreRepository
from app.services.crud_service import CrudService


class ImportacionArchivoService(CrudService[ImportacionArchivo]):
    def __init__(self, conn: asyncpg.Connection) -> None:
        super().__init__(ImportacionArchivoRepository(conn), "ImportacionArchivo")
        self._carrera_repo = CarreraRepository(conn)
        self._estudiante_repo = EstudianteRepository(conn)
        self._inscripcion_repo = InscripcionCuatrimestreRepository(conn)
        self._cursada_repo = CursadaRepository(conn)
        self._parcial_repo = CrudRepository[Parcial](
            conn,
            Parcial,
            CrudTableConfig(
                table_name="parciales",
                columns=("id", "cursada_id", "numero_parcial", "nota", "recuperatorio"),
            ),
        )

    async def importar(
        self,
        file_content: bytes,
        filename: str,
        materia_id: int,
        usuario_id: int,
    ) -> ImportacionArchivo:
        ext = os.path.splitext(filename)[1].lower()
        rows = self._parse(file_content, ext)

        filas_importadas = 0
        filas_errores = 0

        for row in rows:
            try:
                await self._procesar_fila(row, materia_id)
                filas_importadas += 1
            except Exception:
                filas_errores += 1

        result = await self.repo.create(
            usuario_id=usuario_id,
            nombre_archivo=filename,
            filas_importadas=filas_importadas,
            filas_errores=filas_errores,
        )
        return result

    async def listar_por_usuario(self, usuario_id: int) -> list[ImportacionArchivo]:
        return await self.repo.get_by_usuario_id(usuario_id)

    def _parse(self, content: bytes, ext: str) -> list[dict]:
        if ext == ".csv":
            return self._parse_csv(content)
        elif ext in (".xlsx", ".xls"):
            return self._parse_xlsx(content)
        else:
            raise ValueError(f"Formato de archivo no soportado: {ext}")

    def _parse_csv(self, content: bytes) -> list[dict]:
        text = content.decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(text))
        return list(reader)

    def _parse_xlsx(self, content: bytes) -> list[dict]:
        wb = load_workbook(io.BytesIO(content), read_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        headers = [str(h).strip().lower() if h else "" for h in next(rows_iter, [])]
        result = []
        for row in rows_iter:
            result.append(dict(zip(headers, row)))
        wb.close()
        return result

    async def _procesar_fila(self, row: dict, materia_id: int) -> None:
        legajo = str(row.get("legajo", "")).strip()
        if not legajo:
            raise ValueError("legajo vacío")

        apellido, nombre = self._split_apellido_nombre(
            str(row.get("apellido y nombre", "")).strip()
        )
        nota = float(row.get("nota", 0))
        cod_carrera = str(row.get("cod_carrera", "")).strip().upper()
        nro_parcial = int(row.get("nro_parcial", 0))

        carrera = await self._carrera_repo.get_by_codigo(cod_carrera)
        if not carrera:
            raise ValueError(f"Carrera con código '{cod_carrera}' no encontrada")
        carrera_id = carrera.id

        estudiante = await self._estudiante_repo.get_by_legajo_and_carrera(
            legajo, carrera_id
        )
        if not estudiante:
            hoy = date.today()
            estudiante = await self._estudiante_repo.create(
                carrera_id=carrera_id,
                nombre=nombre,
                apellido=apellido,
                legajo=legajo,
                dni=legajo,
                anio_ingreso=hoy.year,
                etapa="temprana",
                porcentaje_carrera=0.0,
                activo=True,
            )
        estudiante_id = estudiante.id

        hoy = date.today()
        cuatrimestre = 1 if hoy.month <= 7 else 2
        anio = hoy.year

        inscripcion = await self._inscripcion_repo.get_by_estudiante_anio_cuatrimestre(
            estudiante_id, anio, cuatrimestre
        )
        if not inscripcion:
            inscripcion = await self._inscripcion_repo.create(
                estudiante_id=estudiante_id,
                anio=anio,
                cuatrimestre=cuatrimestre,
                materilas_anotadas=1,
                activo=True,
            )
        inscripcion_id = inscripcion.id

        cursada = await self._cursada_repo.get_by_estudiante_materia_anio_cuatrimestre(
            estudiante_id, materia_id, anio, cuatrimestre
        )
        if not cursada:
            cursada = await self._cursada_repo.create(
                estudiante_id=estudiante_id,
                materia_id=materia_id,
                inscripcion_id=inscripcion_id,
                anio=anio,
                cuatrimestre=cuatrimestre,
                estado="cursando",
            )
        cursada_id = cursada.id

        await self._parcial_repo.create(
            cursada_id=cursada_id,
            numero_parcial=nro_parcial,
            nota=nota,
            recuperatorio=False,
        )

    def _split_apellido_nombre(self, raw: str) -> tuple[str, str]:
        if ", " in raw:
            parts = raw.split(", ", 1)
            return parts[0].strip(), parts[1].strip()
        return raw.strip(), ""
